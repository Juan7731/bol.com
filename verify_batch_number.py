"""Verify that Batch Number column matches filename"""

import openpyxl
import glob
import os

def verify_batch_numbers():
    """Check if Batch Number column matches filename"""
    print("="*80)
    print("VERIFYING BATCH NUMBER MATCHES FILENAME")
    print("="*80)
    
    # Find all Excel files
    pattern = "batches/**/*.xlsx"
    files = glob.glob(pattern, recursive=True)
    
    if not files:
        print("No Excel files found")
        return
    
    print(f"\nFound {len(files)} Excel file(s)\n")
    
    all_correct = True
    
    for file_path in sorted(files):
        try:
            # Extract batch number from filename (full filename without extension)
            filename = os.path.basename(file_path)
            # Format: S-001.xlsx, SL-001.xlsx, M-001.xlsx
            # Expected batch number should be full filename without extension: "S-001", "SL-001", "M-001"
            expected_batch = filename.replace(".xlsx", "")  # e.g., "S-001", "SL-001", "M-001"
            
            # Open Excel file
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            batch_col_idx = None
            for idx, header in enumerate(headers):
                if header == "Batch Number":
                    batch_col_idx = idx
                    break
            
            if batch_col_idx is None:
                print(f"❌ {filename}: 'Batch Number' column not found")
                all_correct = False
                continue
            
            # Check first data row
            if ws.max_row < 2:
                print(f"⚠️  {filename}: No data rows")
                continue
            
            # Get batch number from first data row
            actual_batch = ws[2][batch_col_idx].value  # Row 2 (first data row)
            actual_batch_str = str(actual_batch) if actual_batch else ""
            
            # Compare
            if actual_batch_str == expected_batch:
                print(f"✅ {filename}: Batch Number column = '{actual_batch_str}' (matches filename)")
            else:
                print(f"❌ {filename}: Batch Number column = '{actual_batch_str}' (expected '{expected_batch}')")
                all_correct = False
            
        except Exception as e:
            print(f"❌ {filename}: Error - {e}")
            all_correct = False
    
    print("\n" + "="*80)
    if all_correct:
        print("✅ ALL FILES: Batch Number matches filename")
    else:
        print("❌ SOME FILES: Batch Number does NOT match filename")
    print("="*80)

if __name__ == "__main__":
    verify_batch_numbers()

