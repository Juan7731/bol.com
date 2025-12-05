"""
Quick test to verify Excel files and system status
"""

import os
import glob
import openpyxl
from order_database import get_processed_count, get_processed_orders_summary

def find_excel_files():
    """Find all Excel files in batches directory"""
    pattern = "batches/**/*.xlsx"
    files = glob.glob(pattern, recursive=True)
    return sorted(files, key=os.path.getmtime, reverse=True)

def test_excel_file(file_path):
    """Test a single Excel file"""
    print(f"\n{'='*80}")
    print(f"Testing: {os.path.basename(file_path)}")
    print(f"Path: {file_path}")
    print('='*80)
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"\n📋 Headers ({len(headers)} columns):")
        for i, header in enumerate(headers, 1):
            print(f"   {chr(64+i)}: {header}")
        
        # Check for required columns (per requirements)
        required = ["Order ID", "Shop", "MP EAN", "Quantity", "Shipping Label", 
                   "Order Time", "Batch Type", "Batch Number", "Order Status"]
        
        missing = [h for h in required if h not in headers]
        if missing:
            print(f"\n❌ Missing columns: {missing}")
        else:
            print(f"\n✅ All required columns present")
        
        # Check for removed column
        if "Customer Name" in headers:
            print(f"❌ 'Customer Name' column still present (should be removed)")
        else:
            print(f"✅ 'Customer Name' column removed")
        
        # Check Shop column
        if "Shop" in headers:
            shop_idx = headers.index("Shop")
            print(f"\n🏪 Shop Column (Column {chr(65+shop_idx)}):")
            
            # Check first 5 data rows
            shop_values = set()
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if row_idx > 6:  # First 5 rows
                    break
                if any(cell.value for cell in row):
                    shop_val = row[shop_idx].value
                    if shop_val:
                        shop_values.add(shop_val)
                        print(f"   Row {row_idx}: {shop_val}")
            
            if shop_values:
                print(f"✅ Shop values found: {sorted(shop_values)}")
            else:
                print(f"⚠️  No shop values found in first rows")
        else:
            print(f"❌ 'Shop' column missing")
        
        # Check ZPL labels
        if "Shipping Label" in headers:
            zpl_idx = headers.index("Shipping Label")
            print(f"\n📦 Shipping Label Column (Column {chr(65+zpl_idx)}):")
            
            zpl_count = 0
            empty_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if row_idx > 10:  # Check first 10 rows
                    break
                if any(cell.value for cell in row):
                    zpl_val = row[zpl_idx].value
                    if zpl_val:
                        zpl_count += 1
                        zpl_str = str(zpl_val)
                        if len(zpl_str) > 50:
                            print(f"   Row {row_idx}: ZPL present ({len(zpl_str)} chars) - {zpl_str[:50]}...")
                        else:
                            print(f"   Row {row_idx}: ZPL present ({len(zpl_str)} chars)")
                    else:
                        empty_count += 1
            
            if zpl_count > 0:
                print(f"✅ ZPL labels found: {zpl_count} rows with labels")
            else:
                print(f"⚠️  No ZPL labels found in first rows ({empty_count} empty)")
        else:
            print(f"❌ 'Shipping Label' column missing")
        
        # Count total rows
        total_rows = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
        print(f"\n📊 Total data rows: {total_rows}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    print("="*80)
    print("QUICK SYSTEM TEST")
    print("="*80)
    
    # Check database
    print("\n📊 Database Status:")
    count = get_processed_count()
    print(f"   Processed orders: {count}")
    summary = get_processed_orders_summary()
    if summary:
        print(f"   By type: {summary}")
    
    # Find Excel files
    print("\n📁 Searching for Excel files...")
    excel_files = find_excel_files()
    
    if not excel_files:
        print("   ⚠️  No Excel files found in batches/ directory")
        print("\n   To generate new files:")
        print("   1. Reset database: Delete bol_orders.db")
        print("   2. Run: python order_processing.py")
        return
    
    print(f"   Found {len(excel_files)} Excel file(s)")
    
    # Test latest file
    if excel_files:
        latest = excel_files[0]
        print(f"\n   Testing latest file: {os.path.basename(latest)}")
        test_excel_file(latest)
    
    # Test all files if multiple
    if len(excel_files) > 1:
        print(f"\n\n{'='*80}")
        print(f"Testing all {len(excel_files)} files...")
        print('='*80)
        for file_path in excel_files[:5]:  # Test first 5
            test_excel_file(file_path)
    
    print("\n" + "="*80)
    print("✅ Test complete!")
    print("="*80)
    print("\nTo test with new orders:")
    print("  1. Delete bol_orders.db to reset processed orders")
    print("  2. Run: python order_processing.py")
    print("  3. Run this test again: python quick_test.py")

if __name__ == "__main__":
    main()

